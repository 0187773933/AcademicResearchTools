#!/usr/bin/env python3

import csv
import json
import os
import random
import re
import sys
import time
from datetime import datetime, timedelta, UTC
from typing import Dict, List, Optional, Set, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import requests
from requests.exceptions import (
	ChunkedEncodingError,
	ConnectionError,
	ReadTimeout,
	RequestException,
)
from tqdm import tqdm
from urllib3.exceptions import ProtocolError

# ============================================================
# CONFIG
# ============================================================

SEARCH_HISTORY_DAYS = ( 365 * 2 )
MAX_REPO_SEARCH_PAGES = 10
MAX_CODE_SEARCH_PAGES = 10
PER_PAGE = 100
MAX_FILES_TO_CHECK = 25
REQUEST_TIMEOUT = 45
MAX_RETRIES = 6
GITHUB_TOKEN = os.environ.get( "GITHUB_TOKEN" , "PUT_TOKEN_HERE" )

if not GITHUB_TOKEN or GITHUB_TOKEN == "PUT_TOKEN_HERE":
	print("Set GITHUB_TOKEN in your environment or edit the script.")
	sys.exit(1)

NOW = datetime.now(UTC)
CUTOFF_DT = NOW - timedelta(days=SEARCH_HISTORY_DAYS)
CUTOFF = CUTOFF_DT.strftime("%Y-%m-%d")

SEARCH_TERMS = [
	"fmri decoder",
	"fmri decoding",
	"fmri classifier",
	"fmri machine learning",
	"fmri deep learning",
	"fmri mvpa",
	"fmri transformer",
	"fmri llm",
	"brain decoding fmri",
	"voxelwise decoding fmri",
	"voxelwise encoding model fmri",
	"fmri pytorch",
	"fmri sklearn",
]

IMPORT_TERMS = [
	"import nilearn",
	"import nibabel",
	"import brainiak",
]

CODE_EXTENSIONS = (
	".py",
	".ipynb",
	".m",
	".r",
	".jl",
)


RESULTS_NDJSON = "fmri_ml_repos.ndjson"
RESULTS_CSV = "fmri_ml_repos.csv"
RESULTS_XLSX = "fmri_ml_repos.xlsx"
STATE_JSON = "fmri_ml_state.json"

# Strong ML / decoder signals
ML_PATTERNS = [
	r"\bfrom\s+sklearn\b",
	r"\bimport\s+sklearn\b",
	r"\bfrom\s+torch\b",
	r"\bimport\s+torch\b",
	r"\bfrom\s+tensorflow\b",
	r"\bimport\s+tensorflow\b",
	r"\bfrom\s+keras\b",
	r"\bimport\s+keras\b",
	r"\bfrom\s+xgboost\b",
	r"\bimport\s+xgboost\b",
	r"\bclassifier\b",
	r"\bclassification\b",
	r"\bcross_val\b",
	r"\bcross_validate\b",
	r"\bcross_val_score\b",
	r"\bconfusion_matrix\b",
	r"\broc_auc\b",
	r"\bfit\s*\(",
	r"\bpredict\s*\(",
	r"\bsearchlight\b",
	r"\bdecoder\b",
	r"\bdecoding\b",
	r"\bmvpa\b",
	r"\bsvm\b",
	r"\blogisticregression\b",
	r"\brandomforestclassifier\b",
	r"\bmlpclassifier\b",
	r"\bcnn\b",
	r"\blstm\b",
	r"\btransformer\b",
]

DOMAIN_PATTERNS = [
	r"\bnilearn\b",
	r"\bnibabel\b",
	r"\bbrainiak\b",
	r"\bfmri\b",
	r"\bbold\b",
	r"\bnifti\b",
	r"\bsearchlight\b",
	r"\bvoxel\b",
	r"\bmasker\b",
	r"\bconfounds\b",
	r"\bbids\b",
]

DECODER_TEXT_TERMS = [
	"decoder",
	"decoding",
	"classifier",
	"classification",
	"mvpa",
	"searchlight",
	"voxelwise",
	"brain decoding",
	"encoding model",
]

ML_REGEXES = [re.compile(p, re.I) for p in ML_PATTERNS]
DOMAIN_REGEXES = [re.compile(p, re.I) for p in DOMAIN_PATTERNS]

HEADERS = {
	"Authorization": f"Bearer {GITHUB_TOKEN}",
	"Accept": "application/vnd.github+json",
	"X-GitHub-Api-Version": "2022-11-28",
	"User-Agent": "fmri-ml-repo-crawler/1.0",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

# ============================================================
# IO / STATE
# ============================================================

# class LiveWriter:
#     def __init__(self, ndjson_path: str, csv_path: str):
#         self.ndjson_path = ndjson_path
#         self.csv_path = csv_path

#         self.ndjson_fp = open(self.ndjson_path, "a", encoding="utf-8")
#         self.csv_fp = open(self.csv_path, "a", newline="", encoding="utf-8")

#         self.csv_writer = csv.DictWriter(
#             self.csv_fp,
#             fieldnames=[
#                 "repo",
#                 "stars",
#                 "forks",
#                 "created_at",
#                 "pushed_at",
#                 "default_branch",
#                 "url",
#                 "description",
#                 "source",
#                 "ml_score",
#                 "matched_files",
#             ],
#         )

#         if self.csv_fp.tell() == 0:
#             self.csv_writer.writeheader()
#             self.csv_fp.flush()

#     def write(self, row: dict):
#         # console
#         print(
#             f'\n{row["stars"]:>6} ⭐  {row["repo"]}\n'
#             f'       {row["url"]}\n'
#             f'       source={row["source"]}  ml_score={row["ml_score"]}\n'
#             f'       files={row["matched_files"]}\n',
#             flush=True
#         )

#         # ndjson
#         self.ndjson_fp.write(json.dumps(row, ensure_ascii=False) + "\n")
#         self.ndjson_fp.flush()

#         # csv
#         self.csv_writer.writerow(row)
#         self.csv_fp.flush()

#     def close(self):
#         self.ndjson_fp.close()
#         self.csv_fp.close()

class LiveWriter:

	def __init__(self, ndjson_path, xlsx_path):

		self.ndjson_fp = open(ndjson_path, "a", encoding="utf-8")

		self.xlsx_path = xlsx_path

		if os.path.exists(xlsx_path):
			self.wb = load_workbook(xlsx_path)
			self.ws = self.wb.active
		else:
			self.wb = Workbook()
			self.ws = self.wb.active

			headers = [
				"repo",
				"stars",
				"forks",
				"created_at",
				"pushed_at",
				"default_branch",
				"url",
				"description",
				"source",
				"ml_score",
				"matched_files",
			]

			self.ws.append(headers)

			for col in range(1, len(headers)+1):
				self.ws.cell(row=1, column=col).font = Font(bold=True)

			self.wb.save(self.xlsx_path)

	def write(self, row):

		print(
			f'\n{row["stars"]:>6} ⭐  {row["repo"]}\n'
			f'       {row["url"]}\n'
			f'       source={row["source"]} ml_score={row["ml_score"]}\n',
			flush=True
		)

		# write ndjson
		self.ndjson_fp.write(json.dumps(row, ensure_ascii=False) + "\n")
		self.ndjson_fp.flush()

		# write xlsx row
		values = [
			row["repo"],
			row["stars"],
			row["forks"],
			row["created_at"],
			row["pushed_at"],
			row["default_branch"],
			row["url"],
			row["description"],
			row["source"],
			row["ml_score"],
			row["matched_files"],
		]

		self.ws.append(values)

		r = self.ws.max_row

		# make URL clickable
		cell = self.ws.cell(row=r, column=7)
		cell.hyperlink = row["url"]
		cell.style = "Hyperlink"

		self.wb.save(self.xlsx_path)

	def close(self):

		self.ndjson_fp.close()
		self.wb.save(self.xlsx_path)


def load_state() -> dict:
	if not os.path.exists(STATE_JSON):
		return {
			"seen_repo_full_names": [],
			"written_repo_full_names": [],
		}
	with open(STATE_JSON, "r", encoding="utf-8") as f:
		return json.load(f)


def save_state(state: dict):
	tmp = STATE_JSON + ".tmp"
	with open(tmp, "w", encoding="utf-8") as f:
		json.dump(state, f, indent=2)
	os.replace(tmp, STATE_JSON)


# ============================================================
# NETWORK / RETRY
# ============================================================

def maybe_sleep_from_rate_limit(resp: requests.Response):
	if resp is None:
		return

	remaining = resp.headers.get("X-RateLimit-Remaining")
	reset = resp.headers.get("X-RateLimit-Reset")

	if remaining == "0" and reset:
		try:
			sleep_for = max(1, int(reset) - int(time.time()) + 2)
			print(f"\nRate limit hit. Sleeping {sleep_for}s...", flush=True)
			time.sleep(sleep_for)
		except Exception:
			time.sleep(30)


def safe_get(url: str, *, params: Optional[dict] = None, stream: bool = False) -> Optional[requests.Response]:
	"""
	Robust GET with retries for chunked transfer weirdness, timeouts, 5xx,
	and secondary rate limiting.
	"""
	last_error = None

	for attempt in range(1, MAX_RETRIES + 1):
		try:
			resp = SESSION.get(
				url,
				params=params,
				timeout=REQUEST_TIMEOUT,
				stream=stream,
			)

			# explicit rate limit / abuse handling
			if resp.status_code in (403, 429):
				body = ""
				try:
					body = resp.text.lower()
				except Exception:
					body = ""

				if (
					"rate limit" in body
					or "secondary rate limit" in body
					or resp.headers.get("X-RateLimit-Remaining") == "0"
				):
					maybe_sleep_from_rate_limit(resp)
					backoff = min(60, 2 ** attempt + random.uniform(0, 1.0))
					print(f"\nRetrying after rate-limit-ish response in {backoff:.1f}s...", flush=True)
					time.sleep(backoff)
					continue

			if resp.status_code >= 500:
				backoff = min(60, 2 ** attempt + random.uniform(0, 1.0))
				print(f"\nServer error {resp.status_code} for {url}. Retry in {backoff:.1f}s...", flush=True)
				time.sleep(backoff)
				continue

			return resp

		except (
			ChunkedEncodingError,
			ConnectionError,
			ReadTimeout,
			ProtocolError,
			RequestException,
		) as e:
			last_error = e
			backoff = min(60, 2 ** attempt + random.uniform(0, 1.0))
			print(f"\nNetwork error on {url}: {type(e).__name__}: {e}")
			print(f"Retry {attempt}/{MAX_RETRIES} in {backoff:.1f}s...", flush=True)
			time.sleep(backoff)

	print(f"\nFailed after retries: {url}")
	if last_error:
		print(f"Last error: {type(last_error).__name__}: {last_error}", flush=True)
	return None


# ============================================================
# GITHUB HELPERS
# ============================================================

def parse_github_datetime(s: str) -> datetime:
	return datetime.fromisoformat(s.replace("Z", "+00:00"))


def repo_is_recent(repo: dict) -> bool:
	created = parse_github_datetime(repo["created_at"])
	pushed = parse_github_datetime(repo["pushed_at"])
	return created >= CUTOFF_DT or pushed >= CUTOFF_DT


def github_search_repositories(query: str, page_position: int = 1) -> List[dict]:
	repos: List[dict] = []

	page_bar = tqdm(
		range(1, MAX_REPO_SEARCH_PAGES + 1),
		desc="Pages",
		leave=False,
		position=page_position,
		dynamic_ncols=True,
	)

	for page in page_bar:
		resp = safe_get(
			"https://api.github.com/search/repositories",
			params={
				"q": query,
				"sort": "stars",
				"order": "desc",
				"per_page": PER_PAGE,
				"page": page,
			},
		)

		if resp is None:
			break

		if resp.status_code != 200:
			print(f"\nRepo search error {resp.status_code}: {resp.text[:400]}", flush=True)
			break

		data = resp.json()
		items = data.get("items", [])
		repos.extend(items)

		page_bar.set_postfix(found=len(repos))

		if len(items) < PER_PAGE:
			break

		time.sleep(0.15)

	return repos


def github_search_code(query: str, page_position: int = 1) -> Set[str]:
	repo_full_names: Set[str] = set()

	page_bar = tqdm(
		range(1, MAX_CODE_SEARCH_PAGES + 1),
		desc="CodePages",
		leave=False,
		position=page_position,
		dynamic_ncols=True,
	)

	for page in page_bar:
		resp = safe_get(
			"https://api.github.com/search/code",
			params={
				"q": query,
				"per_page": PER_PAGE,
				"page": page,
			},
		)

		if resp is None:
			break

		if resp.status_code != 200:
			print(f"\nCode search error {resp.status_code}: {resp.text[:400]}", flush=True)
			break

		data = resp.json()
		items = data.get("items", [])

		for item in items:
			repo = item.get("repository", {})
			full_name = repo.get("full_name")
			if full_name:
				repo_full_names.add(full_name)

		page_bar.set_postfix(found=len(repo_full_names))

		if len(items) < PER_PAGE:
			break

		time.sleep(0.20)

	return repo_full_names


def fetch_repo(full_name: str) -> Optional[dict]:
	resp = safe_get(f"https://api.github.com/repos/{full_name}")
	if resp is None or resp.status_code != 200:
		return None
	return resp.json()


def fetch_readme(owner: str, repo: str) -> str:
	import base64

	resp = safe_get(f"https://api.github.com/repos/{owner}/{repo}/readme")
	if resp is None or resp.status_code != 200:
		return ""

	data = resp.json()
	content = data.get("content", "")
	encoding = data.get("encoding", "")

	if encoding == "base64" and content:
		try:
			return base64.b64decode(content).decode("utf-8", errors="ignore")
		except Exception:
			return ""
	return ""


def fetch_repo_tree(owner: str, repo: str, branch: str) -> List[str]:
	resp = safe_get(
		f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}",
		params={"recursive": "1"},
	)
	if resp is None or resp.status_code != 200:
		return []

	data = resp.json()
	files = []

	for node in data.get("tree", []):
		if node.get("type") != "blob":
			continue
		path = node.get("path", "")
		if path.lower().endswith(CODE_EXTENSIONS):
			files.append(path)

	return files


def fetch_raw_file(owner: str, repo: str, branch: str, path: str) -> str:
	url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{path}"
	resp = safe_get(url)
	if resp is None or resp.status_code != 200:
		return ""
	return resp.text


# ============================================================
# HEURISTICS
# ============================================================

def text_has_decoder_signal(name: str, description: str, readme: str) -> bool:
	blob = f"{name}\n{description}\n{readme}".lower()
	return any(term in blob for term in DECODER_TEXT_TERMS)


def score_code_text(text: str) -> int:
	score = 0
	for rx in ML_REGEXES:
		if rx.search(text):
			score += 1
	for rx in DOMAIN_REGEXES:
		if rx.search(text):
			score += 1
	return score


def choose_candidate_files(paths: List[str]) -> List[str]:
	boosted = []
	fallback = []

	for path in paths:
		low = path.lower()
		if any(k in low for k in [
			"decode", "decoder", "classif", "mvpa", "searchlight",
			"train", "model", "fmri", "bold", "voxel", "nilearn", "brainiak"
		]):
			boosted.append(path)
		else:
			fallback.append(path)

	selected = boosted[:MAX_FILES_TO_CHECK]
	if len(selected) < MAX_FILES_TO_CHECK:
		needed = MAX_FILES_TO_CHECK - len(selected)
		selected.extend(fallback[:needed])

	return selected


def analyze_repo(repo: dict, source: str, file_bar_position: int = 3) -> Optional[dict]:
	owner = repo["owner"]["login"]
	name = repo["name"]
	full_name = repo["full_name"]
	html_url = repo["html_url"]
	description = repo.get("description") or ""
	stars = repo.get("stargazers_count", 0)
	forks = repo.get("forks_count", 0)
	default_branch = repo.get("default_branch") or "HEAD"

	# Must be recent by creation OR push
	if not repo_is_recent(repo):
		return None

	readme = fetch_readme(owner, name)

	# Need at least some domain signal from repo text or imports will rescue it later
	text_blob = f"{name}\n{description}\n{readme}".lower()
	domain_text_ok = any(x in text_blob for x in ["fmri", "bold", "nilearn", "nibabel", "brainiak", "voxel"])

	tree_paths = fetch_repo_tree(owner, name, default_branch)
	if not tree_paths:
		return None

	candidate_paths = choose_candidate_files(tree_paths)
	if not candidate_paths:
		return None

	matched_files = []
	best_score = 0

	file_bar = tqdm(
		candidate_paths,
		desc="Files",
		leave=False,
		position=file_bar_position,
		dynamic_ncols=True,
	)

	for path in file_bar:
		text = fetch_raw_file(owner, name, default_branch, path)
		if not text:
			continue

		score = score_code_text(text)
		if score > best_score:
			best_score = score

		if score >= 4:
			matched_files.append(path)

		file_bar.set_postfix(best=best_score, hits=len(matched_files))

		# early success
		if len(matched_files) >= 3:
			break

		time.sleep(0.05)

	has_decoder_text = text_has_decoder_signal(name, description, readme)

	# Gate logic:
	# - real ML-ish code score required
	# - plus either decoder-ish text or strong domain signals
	if best_score < 4:
		return None

	if not (has_decoder_text or domain_text_ok or source == "import_search"):
		return None

	return {
		"repo": full_name,
		"stars": stars,
		"forks": forks,
		"created_at": repo["created_at"],
		"pushed_at": repo["pushed_at"],
		"default_branch": default_branch,
		"url": html_url,
		"description": description,
		"source": source,
		"ml_score": best_score,
		"matched_files": "; ".join(matched_files[:10]),
	}


# ============================================================
# MAIN
# ============================================================

def main():
	print(f"\nDate cutoff: {CUTOFF}")
	print("Mode: created OR pushed within the past year")
	print("Real-time output: console + NDJSON + XLSX + checkpoint state\n")

	state = load_state()
	seen_repo_full_names: Set[str] = set(state.get("seen_repo_full_names", []))
	written_repo_full_names: Set[str] = set(state.get("written_repo_full_names", []))

	# writer = LiveWriter(RESULTS_NDJSON, RESULTS_CSV)
	writer = LiveWriter(RESULTS_NDJSON, RESULTS_XLSX)

	try:
		# --------------------------------------------
		# Stage 1: keyword repo search
		# --------------------------------------------
		keyword_queries = []
		for term in SEARCH_TERMS:
			keyword_queries.append((f"{term} in:name,description,readme created:>={CUTOFF}", "keyword_created"))
			keyword_queries.append((f"{term} in:name,description,readme pushed:>={CUTOFF}", "keyword_pushed"))

		repo_candidates: Dict[str, Tuple[dict, str]] = {}

		print("Stage 1: keyword repository search\n")

		query_bar = tqdm(
			keyword_queries,
			desc="Queries",
			position=0,
			dynamic_ncols=True,
		)

		for query, source in query_bar:
			short_q = query if len(query) < 90 else query[:87] + "..."
			query_bar.set_postfix(unique=len(repo_candidates), q=short_q)

			repos = github_search_repositories(query, page_position=1)

			for repo in repos:
				full_name = repo["full_name"]
				repo_candidates[full_name] = (repo, source)
				seen_repo_full_names.add(full_name)

			state["seen_repo_full_names"] = sorted(seen_repo_full_names)
			state["written_repo_full_names"] = sorted(written_repo_full_names)
			save_state(state)

		# --------------------------------------------
		# Stage 2: import-based code search
		# --------------------------------------------
		print("\nStage 2: import-based code search\n")

		import_bar = tqdm(
			IMPORT_TERMS,
			desc="ImportQueries",
			position=0,
			dynamic_ncols=True,
		)

		for term in import_bar:
			import_bar.set_postfix(unique=len(repo_candidates), term=term)

			code_query = f'"{term}" in:file language:python'
			full_names = github_search_code(code_query, page_position=1)

			fetch_bar = tqdm(
				sorted(full_names),
				desc="FetchRepos",
				leave=False,
				position=2,
				dynamic_ncols=True,
			)

			for full_name in fetch_bar:
				fetch_bar.set_postfix(repo=full_name)

				if full_name in repo_candidates:
					continue

				repo = fetch_repo(full_name)
				if repo is None:
					continue

				repo_candidates[full_name] = (repo, "import_search")
				seen_repo_full_names.add(full_name)

			state["seen_repo_full_names"] = sorted(seen_repo_full_names)
			state["written_repo_full_names"] = sorted(written_repo_full_names)
			save_state(state)

		# --------------------------------------------
		# Stage 3: analyze repos for actual ML code
		# --------------------------------------------
		repos_to_scan = list(repo_candidates.values())
		repos_to_scan.sort(
			key=lambda item: (
				item[0].get("stargazers_count", 0),
				item[0].get("forks_count", 0),
			),
			reverse=True,
		)

		print(f"\nUnique candidate repos: {len(repos_to_scan)}")
		print("Stage 3: scanning repos for actual fMRI / neuroimaging ML code\n")

		repo_bar = tqdm(
			repos_to_scan,
			desc="Repos",
			position=0,
			dynamic_ncols=True,
		)

		hits = 0

		for repo, source in repo_bar:
			full_name = repo["full_name"]
			repo_bar.set_postfix(hits=hits, repo=full_name)

			if full_name in written_repo_full_names:
				continue

			row = analyze_repo(repo, source, file_bar_position=1)
			if row is None:
				continue

			hits += 1
			writer.write(row)
			written_repo_full_names.add(full_name)

			state["seen_repo_full_names"] = sorted(seen_repo_full_names)
			state["written_repo_full_names"] = sorted(written_repo_full_names)
			save_state(state)

		print("\nDone.")
		print(f"Results saved live to:")
		print(f"  {RESULTS_NDJSON}")
		print(f"  {RESULTS_XLSX}")
		print(f"Checkpoint state:")
		print(f"  {STATE_JSON}")

	finally:
		writer.close()


if __name__ == "__main__":
	main()