#!/usr/bin/env python3

import argparse
import json
import math
from pathlib import Path

import pandas as pd
import requests
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================
# CONFIG
# =============================
CACHE_PATH = Path.home() / ".arxiv-cg" / "cache.jsonl"
ZOTERO_URL = "http://127.0.0.1:9371/exists"
BATCH_SIZE = 50

# =============================
# ARGPARSE
# =============================
parser = argparse.ArgumentParser()
parser.add_argument("--input-csv", help="Optional CSV input instead of cache")
parser.add_argument("--output-prefix", default="arxiv_fmri_filtered")
parser.add_argument("--check-zotero", action="store_true")
parser.add_argument("--only-new", action="store_true", help="Keep only papers not in Zotero")
args = parser.parse_args()

# =============================
# LOAD DATA
# =============================
def load_cache():
    records = []
    if not CACHE_PATH.exists():
        return records

    with open(CACHE_PATH, "r", encoding="utf-8") as f:
        for line in f:
            try:
                records.append(json.loads(line))
            except Exception:
                continue
    return records


if args.input_csv:
    print(f"📄 Loading CSV: {args.input_csv}")
    df = pd.read_csv(args.input_csv)
else:
    print(f"📦 Loading cache: {CACHE_PATH}")
    df = pd.DataFrame(load_cache())

if df.empty:
    print("❌ No data found")
    raise SystemExit(1)

print(f"📊 Total records: {len(df)}")

# =============================
# HELPERS
# =============================
def safe_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
    if pd.isna(value):
        return ""
    return str(value).strip()

def chunks(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i+n]

def col_letter(n):
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

# =============================
# FILTER LOGIC
# =============================
def contains_fmri(row):
    title = safe_text(row.get("title", "")).lower()
    summary = safe_text(row.get("summary", "")).lower()
    return ("fmri" in title) or ("fmri" in summary)

print("🔍 Applying fMRI filter...")
mask = df.apply(contains_fmri, axis=1)
filtered = df[mask].copy()

print(f"✅ Filtered records: {len(filtered)}")

# =============================
# SORT
# =============================
if "published" in filtered.columns:
    filtered["published"] = pd.to_datetime(filtered["published"], errors="coerce", utc=True)
    filtered.sort_values(by="published", ascending=False, inplace=True)

# =============================
# FIX TIMEZONES FOR EXCEL
# =============================
for col in filtered.columns:
    if pd.api.types.is_datetime64_any_dtype(filtered[col]):
        try:
            filtered[col] = filtered[col].dt.tz_localize(None)
        except TypeError:
            pass

# =============================
# CLEAN TEXT FIELDS
# =============================
for col in ["title", "summary", "doi", "pdf_url", "abs_url", "doi_url", "authors", "categories", "id"]:
    if col in filtered.columns:
        filtered[col] = filtered[col].apply(safe_text)

# =============================
# ZOTERO EXISTS CHECK
# =============================
if args.check_zotero:
    print("🧠 Checking against Zotero...")

    filtered = filtered.reset_index(drop=True)
    filtered["in_zotero"] = False

    records = filtered.to_dict("records")
    batched_records = list(chunks(records, BATCH_SIZE))

    for batch_idx, batch in enumerate(tqdm(batched_records, desc="Zotero lookup")):
        payload_queries = []

        for i, r in enumerate(batch):
            payload_queries.append({
                "id": i,
                "title": safe_text(r.get("title", "")),
                # "doi": safe_text(r.get("doi", "")),
            })

        payload = {"queries": payload_queries}

        try:
            resp = requests.post(ZOTERO_URL, json=payload, timeout=15)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            print(f"⚠️ Zotero request failed for batch {batch_idx + 1}: {e}")
            continue

        batch_start = batch_idx * BATCH_SIZE

        for result in data.get("results", []):
            local_id = result.get("id")
            exists = bool(result.get("exists", False))

            if isinstance(local_id, int):
                global_idx = batch_start + local_id
                if 0 <= global_idx < len(filtered):
                    filtered.at[global_idx, "in_zotero"] = exists

    hits = int(filtered["in_zotero"].sum())
    print(f"📊 Zotero hits: {hits} / {len(filtered)}")

    if args.only_new:
        filtered = filtered[~filtered["in_zotero"]].copy()
        print(f"🆕 Remaining (not in Zotero): {len(filtered)}")

# =============================
# OPTIONAL NICE DATE FORMAT
# =============================
if "published" in filtered.columns:
    filtered["published"] = filtered["published"].dt.strftime("%Y-%m-%d")

# =============================
# SAVE CSV
# =============================
csv_path = f"{args.output_prefix}.csv"
filtered.to_csv(csv_path, index=False)
print(f"📄 Saved CSV: {csv_path}")

# =============================
# SAVE XLSX
# =============================
xlsx_path = f"{args.output_prefix}.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "filtered"

for r in dataframe_to_rows(filtered, index=False, header=True):
    ws.append(r)

header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

for field in ["pdf_url", "abs_url", "doi_url"]:
    if field in header_map:
        col = col_letter(header_map[field])
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col}{row}"]
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(xlsx_path)
print(f"📊 Saved XLSX: {xlsx_path}")