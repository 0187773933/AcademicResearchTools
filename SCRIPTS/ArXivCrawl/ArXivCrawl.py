#!/usr/bin/env python3

import argparse
import json
import random
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
import xml.etree.ElementTree as ET

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================================
# CONFIG
# ============================================================

BASE_URL = "http://export.arxiv.org/api/query"

DEFAULT_DAYS_BACK = 365
DEFAULT_PAGE_SIZE = 200
DEFAULT_TIMEOUT = 60
DEFAULT_DELAY_BASE = 3.0
DEFAULT_DELAY_JITTER = 1.5
DEFAULT_MAX_RETRIES = 5

CACHE_DIR = Path.home() / ".arxiv-cg"
CACHE_FILE = CACHE_DIR / "cache.jsonl"
SEEN_FILE = CACHE_DIR / "seen_ids.txt"

OUTPUT_PREFIX = "arxiv_fmri_ml"

SEARCH_TERMS = [
    "fmri decoder",
    "fmri decoding",
    "fmri classifier",
    "fmri machine learning",
    "fmri deep learning",
    "fmri transformer",
    "fmri llm",
    "brain decoding fmri",
    "voxelwise decoding fmri",
    "voxelwise encoding model fmri",
    "inner speech fmri",
    "covert speech fmri",
    "imagined speech fmri",
    "silent speech decoding",
    "neural speech decoding fmri",
    "brain computer interface fmri speech",
    "semantic decoding fmri",
    "language decoding fmri",
    "thought decoding fmri",
    "multivoxel pattern analysis fmri",
    "mvpa fmri decoding",
    "encoding model fmri",
    "decoding model fmri",
]

CATEGORIES = ["q-bio.NC", "cs.LG", "stat.ML"]

NS = {
    "atom": "http://www.w3.org/2005/Atom",
    "opensearch": "http://a9.com/-/spec/opensearch/1.1/",
    "arxiv": "http://arxiv.org/schemas/atom",
}


# ============================================================
# ARGPARSE
# ============================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Crawl arXiv for recent fMRI / decoding related papers with local cache."
    )
    parser.add_argument(
        "--days-back",
        type=int,
        default=DEFAULT_DAYS_BACK,
        help=f"Only include papers published within the last N days (default: {DEFAULT_DAYS_BACK}).",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=DEFAULT_PAGE_SIZE,
        help=f"Results per API page (default: {DEFAULT_PAGE_SIZE}, arXiv allows up to 2000).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=DEFAULT_TIMEOUT,
        help=f"HTTP timeout in seconds (default: {DEFAULT_TIMEOUT}).",
    )
    parser.add_argument(
        "--output-prefix",
        default=OUTPUT_PREFIX,
        help=f"Output prefix for CSV/XLSX (default: {OUTPUT_PREFIX}).",
    )
    parser.add_argument(
        "--rebuild-output-from-cache",
        action="store_true",
        help="Do not crawl; just rebuild CSV/XLSX from existing cache.",
    )
    parser.add_argument(
        "--clear-cache",
        action="store_true",
        help="Delete ~/.arxiv-cg cache files before running.",
    )
    return parser.parse_args()


# ============================================================
# CACHE HELPERS
# ============================================================

def ensure_cache_dir() -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)


def clear_cache() -> None:
    if CACHE_FILE.exists():
        CACHE_FILE.unlink()
    if SEEN_FILE.exists():
        SEEN_FILE.unlink()


def load_seen_ids() -> Set[str]:
    if not SEEN_FILE.exists():
        return set()
    seen: Set[str] = set()
    with open(SEEN_FILE, "r", encoding="utf-8") as f:
        for line in f:
            x = line.strip()
            if x:
                seen.add(x)
    return seen


def append_seen_id(arxiv_id: str) -> None:
    with open(SEEN_FILE, "a", encoding="utf-8") as f:
        f.write(arxiv_id + "\n")


def load_cache_records() -> List[Dict]:
    if not CACHE_FILE.exists():
        return []

    records: List[Dict] = []
    with open(CACHE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return records


def append_cache_record(record: Dict) -> None:
    with open(CACHE_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


# ============================================================
# NETWORK
# ============================================================

def build_session(max_retries: int) -> requests.Session:
    session = requests.Session()

    retry = Retry(
        total=max_retries,
        read=max_retries,
        connect=max_retries,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
        respect_retry_after_header=True,
    )

    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)

    session.headers.update(
        {
            "User-Agent": "arxiv-cg/1.0 (requests; local cache; contact: local-script)"
        }
    )

    return session


def polite_sleep(delay_base: float, delay_jitter: float) -> None:
    time.sleep(delay_base + random.uniform(0, delay_jitter))


def fetch_page(
    session: requests.Session,
    params: Dict[str, str],
    timeout: int,
    delay_base: float,
    delay_jitter: float,
) -> Optional[str]:
    try:
        response = session.get(BASE_URL, params=params, timeout=timeout)
    except requests.RequestException as e:
        print(f"⚠️ Request error: {e}")
        polite_sleep(delay_base, delay_jitter)
        return None

    if response.status_code != 200:
        print(f"⚠️ HTTP {response.status_code} for params start={params.get('start')}")
        polite_sleep(delay_base, delay_jitter)
        return None

    text = response.text
    if not text.strip():
        print(f"⚠️ Empty response for params start={params.get('start')}")
        polite_sleep(delay_base, delay_jitter)
        return None

    return text


# ============================================================
# XML PARSING
# ============================================================

def safe_parse_xml(xml_text: str) -> Optional[ET.Element]:
    try:
        return ET.fromstring(xml_text)
    except ET.ParseError as e:
        print(f"⚠️ XML parse error: {e}")
        return None


def parse_date(date_str: str) -> Optional[datetime]:
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
    except Exception:
        return None


def get_text(parent: ET.Element, path: str) -> str:
    el = parent.find(path, NS)
    if el is None or el.text is None:
        return ""
    return el.text.strip()


def parse_entry(entry: ET.Element) -> Dict:
    entry_id = get_text(entry, "atom:id")
    title = get_text(entry, "atom:title")
    summary = get_text(entry, "atom:summary").replace("\n", " ").strip()
    published = get_text(entry, "atom:published")
    updated = get_text(entry, "atom:updated")

    authors: List[str] = []
    for author in entry.findall("atom:author", NS):
        name = get_text(author, "atom:name")
        if name:
            authors.append(name)

    categories: List[str] = []
    for cat in entry.findall("atom:category", NS):
        term = cat.attrib.get("term", "").strip()
        if term:
            categories.append(term)

    pdf_url = ""
    for link in entry.findall("atom:link", NS):
        if link.attrib.get("title") == "pdf":
            pdf_url = link.attrib.get("href", "").strip()
            break

    doi = get_text(entry, "arxiv:doi")

    primary_category_el = entry.find("arxiv:primary_category", NS)
    primary_category = ""
    if primary_category_el is not None:
        primary_category = primary_category_el.attrib.get("term", "").strip()

    return {
        "id": entry_id,
        "title": title,
        "authors": ", ".join(authors),
        "published": published,
        "updated": updated,
        "summary": summary,
        "categories": ", ".join(categories),
        "primary_category": primary_category,
        "pdf_url": pdf_url,
        "abs_url": entry_id,
        "doi": doi,
        "doi_url": f"https://doi.org/{doi}" if doi else "",
    }


def get_total_results(root: ET.Element) -> int:
    total_el = root.find("opensearch:totalResults", NS)
    if total_el is None or total_el.text is None:
        return 0
    try:
        return int(total_el.text.strip())
    except ValueError:
        return 0


def get_entries(root: ET.Element) -> List[ET.Element]:
    return root.findall("atom:entry", NS)


# ============================================================
# OUTPUT
# ============================================================

def records_to_dataframe(records: List[Dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(
            columns=[
                "id",
                "title",
                "authors",
                "published",
                "updated",
                "summary",
                "categories",
                "primary_category",
                "pdf_url",
                "abs_url",
                "doi",
                "doi_url",
            ]
        )

    df = pd.DataFrame(records)

    if "published" in df.columns:
        df["published_dt"] = pd.to_datetime(df["published"], errors="coerce", utc=True)
    else:
        df["published_dt"] = pd.NaT

    if "updated" in df.columns:
        df["updated_dt"] = pd.to_datetime(df["updated"], errors="coerce", utc=True)
    else:
        df["updated_dt"] = pd.NaT

    df.sort_values(
        by=["published_dt", "updated_dt", "title"],
        ascending=[False, False, True],
        inplace=True,
        na_position="last",
    )

    df.drop(columns=["published_dt", "updated_dt"], inplace=True, errors="ignore")
    df.reset_index(drop=True, inplace=True)

    return df


def save_csv(df: pd.DataFrame, output_prefix: str) -> Path:
    csv_path = Path(f"{output_prefix}.csv")
    df.to_csv(csv_path, index=False)
    return csv_path


def save_xlsx(df: pd.DataFrame, output_prefix: str) -> Path:
    xlsx_path = Path(f"{output_prefix}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "arxiv_results"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    header_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}

    def col_letter(col_num: int) -> str:
        result = ""
        while col_num:
            col_num, rem = divmod(col_num - 1, 26)
            result = chr(65 + rem) + result
        return result

    for field in ["pdf_url", "abs_url", "doi_url"]:
        if field in header_index:
            col_num = header_index[field]
            letter = col_letter(col_num)
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{letter}{row}"]
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

    wb.save(xlsx_path)
    return xlsx_path


# ============================================================
# QUERY BUILDING
# ============================================================

def build_query(term: str, categories: List[str]) -> str:
    category_query = " OR ".join(f"cat:{c}" for c in categories)
    return f"({term}) AND ({category_query})"


# ============================================================
# CRAWL
# ============================================================

def crawl(
    days_back: int,
    page_size: int,
    timeout: int,
    output_prefix: str,
) -> Tuple[int, Path, Path]:
    ensure_cache_dir()

    cutoff_date = datetime.now(timezone.utc) - timedelta(days=days_back)

    print("📦 Loading cache...")
    seen_ids = load_seen_ids()
    print(f"🧠 Seen IDs: {len(seen_ids)}")

    session = build_session(DEFAULT_MAX_RETRIES)

    new_records_count = 0

    for term in SEARCH_TERMS:
        print(f"\n🔎 TERM: {term}")

        query = build_query(term, CATEGORIES)
        start = 0
        total_results: Optional[int] = None
        term_new = 0
        term_seen = 0
        term_old = 0
        term_failed_pages = 0

        pbar: Optional[tqdm] = None

        while True:
            params = {
                "search_query": query,
                "start": str(start),
                "max_results": str(page_size),
                "sortBy": "submittedDate",
                "sortOrder": "descending",
            }

            xml_text = fetch_page(
                session=session,
                params=params,
                timeout=timeout,
                delay_base=DEFAULT_DELAY_BASE,
                delay_jitter=DEFAULT_DELAY_JITTER,
            )

            if xml_text is None:
                term_failed_pages += 1
                print("⚠️ Skipping failed page and moving to next term.")
                break

            root = safe_parse_xml(xml_text)
            if root is None:
                term_failed_pages += 1
                print("⚠️ Could not parse page and moving to next term.")
                break

            if total_results is None:
                total_results = get_total_results(root)
                print(f"📊 total results: {total_results}")
                pbar = tqdm(
                    total=total_results,
                    desc=term,
                    unit="papers",
                    dynamic_ncols=True,
                    smoothing=0.1,
                )

            entries = get_entries(root)
            if not entries:
                print("ℹ️ No entries on this page.")
                break

            stop_term = False

            for entry in entries:
                record = parse_entry(entry)

                entry_id = record.get("id", "").strip()
                if not entry_id:
                    if pbar:
                        pbar.update(1)
                    continue

                published_dt = parse_date(record.get("published", ""))
                if published_dt is None:
                    if pbar:
                        pbar.update(1)
                    continue

                if published_dt < cutoff_date:
                    term_old += 1
                    stop_term = True
                    if pbar:
                        pbar.update(1)
                    break

                if entry_id in seen_ids:
                    term_seen += 1
                    if pbar:
                        pbar.update(1)
                    continue

                seen_ids.add(entry_id)
                append_seen_id(entry_id)
                append_cache_record(record)

                term_new += 1
                new_records_count += 1

                if pbar:
                    pbar.update(1)

            if stop_term:
                print("🛑 cutoff reached, stopping this term early")
                break

            start += page_size

            if total_results is not None and start >= total_results:
                break

            polite_sleep(DEFAULT_DELAY_BASE, DEFAULT_DELAY_JITTER)

        if pbar:
            pbar.close()

        print(
            f"✅ term done | new={term_new} | seen={term_seen} | old_cutoff_hits={term_old} | failed_pages={term_failed_pages}"
        )

    print(f"\n✅ New papers this run: {new_records_count}")

    print("📚 Loading full cache for output...")
    all_records = load_cache_records()
    df = records_to_dataframe(all_records)

    csv_path = save_csv(df, output_prefix)
    xlsx_path = save_xlsx(df, output_prefix)

    print(f"📄 CSV saved: {csv_path}")
    print(f"📊 XLSX saved: {xlsx_path}")

    return new_records_count, csv_path, xlsx_path


# ============================================================
# REBUILD OUTPUT
# ============================================================

def rebuild_output_only(output_prefix: str) -> Tuple[Path, Path]:
    ensure_cache_dir()
    print("📚 Rebuilding outputs from existing cache...")
    records = load_cache_records()
    df = records_to_dataframe(records)
    csv_path = save_csv(df, output_prefix)
    xlsx_path = save_xlsx(df, output_prefix)
    print(f"📄 CSV saved: {csv_path}")
    print(f"📊 XLSX saved: {xlsx_path}")
    return csv_path, xlsx_path


# ============================================================
# MAIN
# ============================================================

def main() -> int:
    args = parse_args()

    ensure_cache_dir()

    if args.clear_cache:
        print("🧹 Clearing cache...")
        clear_cache()

    if args.rebuild_output_from_cache:
        rebuild_output_only(args.output_prefix)
        return 0

    try:
        crawl(
            days_back=args.days_back,
            page_size=args.page_size,
            timeout=args.timeout,
            output_prefix=args.output_prefix,
        )
        return 0
    except KeyboardInterrupt:
        print("\n⚠️ Interrupted by user.")
        return 130
    except Exception as e:
        print(f"\n❌ Fatal error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())