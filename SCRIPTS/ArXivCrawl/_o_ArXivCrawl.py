#!/usr/bin/env python3

import requests
import time
import random
import pandas as pd
from datetime import datetime, timedelta, timezone
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================
# CONFIG
# =============================
BASE_URL = "http://export.arxiv.org/api/query"

DAYS_BACK = 365
PAGE_SIZE = 200  # safe chunk size (<=2000 allowed, but smaller is safer)
MAX_RETRIES = 5

DELAY_BASE = 3.0  # arXiv recommends 3 seconds
DELAY_JITTER = 1.5

OUTPUT_PREFIX = "arxiv_fmri_ml"

SEARCH_TERMS = [
    "fmri decoder",
    "fmri decoding",
    "fmri classifier",
    "fmri machine learning",
    "fmri deep learning",
    "fmri transformer",
    "fmri llm",
    "brain decoding fmri",
    "voxelwise decoding fmri",
    "voxelwise encoding model fmri",

    "inner speech fmri",
    "covert speech fmri",
    "imagined speech fmri",
    "silent speech decoding",
    "neural speech decoding fmri",
    "brain computer interface fmri speech",
    "semantic decoding fmri",
    "language decoding fmri",
    "thought decoding fmri",

    "foundation model brain decoding",
    "transformer brain encoding",
    "multivoxel pattern analysis fmri",
    "mvpa fmri decoding",
    "encoding model fmri",
    "decoding model fmri",
]

CATEGORIES = ["q-bio.NC", "cs.LG", "stat.ML"]

# =============================
# DATE FILTER
# =============================
cutoff_date = datetime.now(timezone.utc) - timedelta(days=DAYS_BACK)

# =============================
# HELPERS
# =============================
def sleep_with_jitter():
    delay = DELAY_BASE + random.uniform(0, DELAY_JITTER)
    time.sleep(delay)

def fetch_with_retry(params):
    for attempt in range(MAX_RETRIES):
        try:
            response = requests.get(BASE_URL, params=params, timeout=60)
            if response.status_code == 200:
                return response.text
            else:
                print(f"⚠️ HTTP {response.status_code}")
        except requests.RequestException as e:
            print(f"⚠️ Request error: {e}")

        wait = (2 ** attempt) + random.uniform(0, 1)
        print(f"⏳ retrying in {wait:.1f}s...")
        time.sleep(wait)

    raise RuntimeError("❌ Failed after retries")

def parse_entry(entry):
    ns = {"atom": "http://www.w3.org/2005/Atom"}

    def get_text(tag):
        el = entry.find(f"atom:{tag}", ns)
        return el.text.strip() if el is not None else ""

    title = get_text("title")
    summary = get_text("summary")
    published = get_text("published")
    updated = get_text("updated")
    entry_id = get_text("id")

    authors = [
        a.find("atom:name", ns).text
        for a in entry.findall("atom:author", ns)
    ]

    categories = [
        c.attrib.get("term")
        for c in entry.findall("atom:category", ns)
    ]

    pdf_url = ""
    for link in entry.findall("atom:link", ns):
        if link.attrib.get("title") == "pdf":
            pdf_url = link.attrib.get("href")

    return {
        "id": entry_id,
        "title": title,
        "authors": ", ".join(authors),
        "published": published,
        "updated": updated,
        "summary": summary.replace("\n", " "),
        "categories": ", ".join(categories),
        "pdf_url": pdf_url,
        "abs_url": entry_id,
    }

# =============================
# MAIN SEARCH LOOP
# =============================
seen_ids = set()
records = []

for term in SEARCH_TERMS:
    print(f"\n🔎 TERM: {term}")

    category_query = " OR ".join([f"cat:{c}" for c in CATEGORIES])
    full_query = f"({term}) AND ({category_query})"

    start = 0
    total_results = None

    while True:
        params = {
            "search_query": full_query,
            "start": start,
            "max_results": PAGE_SIZE,
            "sortBy": "submittedDate",
            "sortOrder": "descending",
        }

        print(f"➡️ start={start}")

        xml_text = fetch_with_retry(params)

        root = ET.fromstring(xml_text)
        ns = {
            "atom": "http://www.w3.org/2005/Atom",
            "opensearch": "http://a9.com/-/spec/opensearch/1.1/",
        }

        if total_results is None:
            total_el = root.find("opensearch:totalResults", ns)
            total_results = int(total_el.text)
            print(f"📊 total results: {total_results}")

        entries = root.findall("atom:entry", ns)
        if not entries:
            break

        for entry in entries:
            data = parse_entry(entry)

            if data["id"] in seen_ids:
                continue

            try:
                pub_date = datetime.fromisoformat(
                    data["published"].replace("Z", "+00:00")
                )
            except:
                continue

            if pub_date < cutoff_date:
                continue

            seen_ids.add(data["id"])
            records.append(data)

        start += PAGE_SIZE

        if start >= total_results:
            break

        sleep_with_jitter()

print(f"\n✅ Total unique papers: {len(records)}")

# =============================
# DATAFRAME
# =============================
df = pd.DataFrame(records)

df["published"] = pd.to_datetime(df["published"])
df.sort_values(by="published", ascending=False, inplace=True)

# =============================
# SAVE CSV
# =============================
csv_path = f"{OUTPUT_PREFIX}.csv"
df.to_csv(csv_path, index=False)
print(f"📄 CSV saved: {csv_path}")

# =============================
# SAVE XLSX (CLICKABLE)
# =============================
xlsx_path = f"{OUTPUT_PREFIX}.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "arxiv_results"

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

for row in range(2, ws.max_row + 1):
    pdf_cell = ws[f"G{row}"]
    abs_cell = ws[f"H{row}"]

    pdf_cell.hyperlink = pdf_cell.value
    pdf_cell.style = "Hyperlink"

    abs_cell.hyperlink = abs_cell.value
    abs_cell.style = "Hyperlink"

wb.save(xlsx_path)

print(f"📊 XLSX saved: {xlsx_path}")